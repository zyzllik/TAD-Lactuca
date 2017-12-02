# -*- coding:utf-8 -*-
"""
@author: loopgan
@time: 9/8/17 5:29 PM
"""
from src.utils import load_data
import os
import matplotlib.pyplot as plt
import datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook

from sklearn.preprocessing import MinMaxScaler
from matplotlib.ticker import MultipleLocator
from matplotlib.ticker import FormatStrFormatter


def get_interval_different_resolution(interval_dir, bin_size, bin_number):
    '''
    :param interval_dir: 计算出的位点区间存放的目录，default = $root_path/$interval_dir
    :param bin_number: 位点上下游bin的个数，default = 10
    :param bin_size: bin的大小，default = 40kb
    :return: 2208个boundary,2208个non-boundary,文件命名为tad_b_$num.bed或tad_not_b_$num.bed
    '''
    with open("%s/cache/tad_boundary_center.txt" % root_path, "w") as r:
        with open("%s/data/tad_b_2012_IMR90_hg18.txt" % root_path) as f:
            for line in f.readlines():
                if line.split()[1] == line.split()[2]:
                    r.write(line.split()[0] + "\t" + str(int(line.split()[1]) - int(bin_size / 2)) + "\n")
                else:
                    tmp_center = str(int((int(line.split()[1]) + int(line.split()[2])) / 2) - int(bin_size / 2))
                    r.write(line.split()[0] + "\t" + tmp_center + "\n")
    with open("%s/cache/tad_not_boundary_center.txt" % root_path, "w") as r:
        with open("%s/data/input_IMR90_TADBoundries.txt" % root_path) as f:
            for line in f.readlines():
                if line.split()[-1] == "NB":
                    r.write(line.split()[0] + "\t" + str(int(line.split()[1]) + 50000 - int(bin_size / 2)) + "\n")
    with open("%s/cache/tad_boundary_center.txt" % root_path) as f:
        line_num = 1
        for line in f.readlines():
            with open("%s/tad_b_%s.bed" % (interval_dir, line_num), "w") as r:
                for i in range(-bin_number, bin_number + 1):
                    r.write(line.split()[0] + "\t" + str(int(line.split()[1]) + i * bin_size) + "\t" + str(
                        int(line.split()[1]) + (i + 1) * bin_size) + "\n")
            line_num += 1
    with open("%s/cache/tad_not_boundary_center.txt" % root_path) as f:
        line_num = 1
        for line in f.readlines():
            with open("%s/tad_not_b_%s.bed" % (interval_dir, line_num), "w") as r:
                for i in range(-bin_number, bin_number + 1):
                    tmp_start = int(line.split()[1]) + i * bin_size if int(
                        int(line.split()[1]) + i * bin_size) > 0 else 0
                    tmp_end = int(line.split()[1]) + (i + 1) * bin_size if int(
                        int(line.split()[1]) + (i + 1) * bin_size) > 0 else 0
                    r.write(line.split()[0] + "\t" + str(tmp_start) + "\t" + str(tmp_end) + "\n")
            line_num += 1


def get_interval_signal(interval_dir, cell_type_dir):
    """
    用bwtool summary 计算间隔的信号强度
    :param interval_dir:interval的路径
    :param cell_type_dir:组蛋白的路径
    :return:在interval路径下生产以目录为区分的个特征信号文件，文件命名为tad_b_$num.signal或tad_not_b_$num.signal
    """
    for file in os.listdir(interval_dir):
        for file_histone in os.listdir(cell_type_dir):
            if not os.path.exists(interval_dir + file_histone.split('.')[0]):
                os.mkdir(interval_dir + file_histone.split('.')[0])
            if len(os.listdir(interval_dir + file_histone.split('.')[0])) != 2208 * 2:
                if file.startswith('tad_b') and file_histone.endswith('.bigwig'):
                    os.system('/opt/bwtool/bin/bwtool summary %s %s %s -with-sum -fill=0' % (
                        interval_dir + file, cell_type_dir + file_histone,
                        interval_dir + file_histone.split('.')[0] + '/' +
                        file.split('.bed')[0] + '.signal'))
                if file.startswith('tad_not_b') and file_histone.endswith('.bigwig'):
                    os.system('/opt/bwtool/bin/bwtool summary %s %s %s -with-sum -fill=0' % (
                        interval_dir + file, cell_type_dir + file_histone,
                        interval_dir + file_histone.split('.')[0] + '/' +
                        file.split('.bed')[0] + '.signal'))
            else:
                print("The signal have been calculated, %s will be passed" % get_interval_signal.__name__.upper())
                return 0


def write_to_xlsx(signal_file_dir, feature_file):
    """
    :param signal_file_dir: 信号文件目录
    :param feature_file: 返回文件的名称
    :return: 某细胞系下的信号文件，excel形式
    """
    if os.path.exists(feature_file):
        print("The file already exists, %s will be pass" % write_to_xlsx.__name__.upper())
        return 0
    workbook = Workbook()
    worksheet_y = workbook.create_sheet("y")
    worksheet_n = workbook.create_sheet("n")
    workbook.remove_sheet(workbook.get_sheet_by_name("Sheet"))
    histone_type = []
    for file in os.listdir(signal_file_dir):
        if os.path.isdir(signal_file_dir + file):
            histone_type.append(file)
    y_file = ["tad_b_%s.signal" % i for i in range(1, 2209)]
    for histone in histone_type:
        tmp_index = histone_type.index(histone)
        shitf = 4 + tmp_index * (bin_number * 2 + 1) + 1
        for i in range(-bin_number, bin_number + 1):
            worksheet_y.cell(row=1, column=shitf, value=histone + "_" + str(i))
            worksheet_n.cell(row=1, column=shitf, value=histone + "_" + str(i))
            shitf += 1
        for file in y_file:
            with open(signal_file_dir + histone + "/" + file) as f:
                tmp_line_num = 1
                for line in f.readlines():
                    tmp_row = int(file.split(".")[0].split("_")[-1]) + 1
                    tmp_col = int(4 + histone_type.index(histone) * (bin_number * 2 + 1) + tmp_line_num)
                    tmp_line_num += 1
                    tmp_value = float(line.split()[9]) if line.split()[9] != "NA" else 0
                    worksheet_y.cell(row=tmp_row, column=tmp_col, value=tmp_value)
                    if histone_type.index(histone) == 0:
                        worksheet_y.cell(row=tmp_row, column=4, value=1)
                        worksheet_y.cell(row=1, column=1, value="chr#")
                        worksheet_y.cell(row=1, column=2, value="start")
                        worksheet_y.cell(row=1, column=3, value="end")
                        worksheet_y.cell(row=1, column=4, value="label")
                        for i in range(1, 4):
                            tmp_value = int(line.split()[i - 1]) if not line.split()[i - 1].startswith("chr") else \
                                line.split()[i - 1]
                            worksheet_y.cell(row=tmp_row, column=i, value=tmp_value)
            with open(signal_file_dir + histone + "/" + file.replace("tad_b", "tad_not_b")) as f:
                tmp_line_num = 1
                for line in f.readlines():
                    tmp_row = int(file.split(".")[0].split("_")[-1]) + 1
                    tmp_col = int(4 + histone_type.index(histone) * (bin_number * 2 + 1) + tmp_line_num)
                    tmp_line_num += 1
                    tmp_value = float(line.split()[9]) if line.split()[9] != "NA" else 0
                    worksheet_n.cell(row=tmp_row, column=tmp_col, value=tmp_value)
                    if histone_type.index(histone) == 0:
                        worksheet_n.cell(row=tmp_row, column=4, value=0)
                        worksheet_n.cell(row=1, column=1, value="chr#")
                        worksheet_n.cell(row=1, column=2, value="start")
                        worksheet_n.cell(row=1, column=3, value="end")
                        worksheet_n.cell(row=1, column=4, value="label")
                        for i in range(1, 4):
                            tmp_value = int(line.split()[i - 1]) if not line.split()[i - 1].startswith("chr") else \
                                line.split()[i - 1]
                            worksheet_n.cell(row=tmp_row, column=i, value=tmp_value)
    workbook.save(feature_file)


def plot_dis(feature_file, label_file, cell_line, feature_dimensional):
    '''
    :param dir_interval: 各信号所在目录
    :param cell_type: 组蛋白的类型
    :return: 在pic/box目录返回各种信号正负样本的余弦相似度，以箱线图形式。
    '''
    if os.path.isfile(feature_file):
        print("The file already exists, %s will be passed" % plot_dis.__name__.upper())
        # return
    x = [i for i in range(-bin_number, bin_number + 1)]
    data, target = load_data.data(feature=feature_file)
    label = load_data.get_label(data=label_file)
    fig, axes = plt.subplots(nrows=3, ncols=3, sharex=True, sharey=True)
    fig_1, axes_1 = plt.subplots(nrows=3, ncols=3, sharex=True)
    i_0 = 0
    legend_0 = 0
    for row in axes:
        for col in row:
            tad_b = np.mean(data[0:tad_num, i_0 * feature_dimensional:(i_0 + 1) * feature_dimensional], axis=0)
            not_tad_b = np.mean(data[tad_num:data.shape[0], i_0 * feature_dimensional:(i_0 + 1) * feature_dimensional],
                                axis=0)
            tad_b = MinMaxScaler().fit_transform(tad_b.reshape(-1, 1))
            not_tad_b = MinMaxScaler().fit_transform(not_tad_b.reshape(-1, 1))
            if not legend_0:
                y, = col.plot(x, tad_b, label='b')
                n, = col.plot(x, not_tad_b, label='not b')
                legend_0 = 1
            else:
                col.plot(x, tad_b, label='b')
                col.plot(x, not_tad_b, label='not b')
            col.set_title(label[i_0], size=8)
            col.set_ylim([-0.05, 1.05])
            col.plot([0, 0], [0, 1.05], linestyle='--')
            i_0 += 1
    i_1 = 0
    legend_1 = 0
    for row in axes_1:
        for col in row:
            tad_b = np.mean(data[0:tad_num, i_1 * feature_dimensional:(i_1 + 1) * feature_dimensional], axis=0).reshape(
                -1, 1) / 10000
            not_tad_b = np.mean(data[tad_num:data.shape[0], i_1 * feature_dimensional:(i_1 + 1) * feature_dimensional],
                                axis=0).reshape(-1, 1) / 10000
            if not legend_1:
                y_1, = col.plot(x, tad_b, label='b')
                n_1, = col.plot(x, not_tad_b, label='not b')
                legend_1 = 1
            else:
                col.plot(x, tad_b, label='b')
                col.plot(x, not_tad_b, label='not b')
            col.set_title(label[i_1], size=8)
            ymajorLocator = MultipleLocator(0.5)
            ymajorFormatter = FormatStrFormatter('%1.1f')
            yminorLocator = MultipleLocator(0.1)
            col.yaxis.set_major_locator(ymajorLocator)
            col.yaxis.set_major_formatter(ymajorFormatter)
            col.yaxis.set_minor_locator(yminorLocator)
            i_1 += 1
    fig.text(0.5, 0.03, 'relative distance from center', ha='center', size=10)
    fig.text(0.04, 0.5, 'normalized signal', va='center', rotation='vertical', size=10)
    fig.legend([y, n], ['Boundary', 'Non-Boundary'], prop={'size': '7'})
    fig.savefig("%s/dis_all_feature/%s.eps" % (pic_path, bin_number), format='eps', dpi=dpi)

    # fig_1.tight_layout()
    fig_1.subplots_adjust(wspace=0.3)
    fig_1.text(0.5, 0.04, 'relative distance from center(bin)', ha='center', size=10)
    fig_1.text(0.04, 0.5, 'signal (scale 1:%s)' % (r'$10^4$'), va='center', rotation='vertical', size=10)
    fig_1.legend([y_1, n_1], ['Boundary', 'Non-Boundary'], prop={'size': '7'})
    fig_1.savefig("%s/dis_all_feature/%s_not_nor.eps" % (pic_path, bin_number), format='eps', dpi=dpi)


def plot_corr(feature_file, label_file, cell_line):
    '''
    计算cosine similarity并画相形图
    :param feature_file:
    :param label_file:
    :param cell_line:
    :return:
    '''
    if os.path.isfile(feature_file):
        print("The file already exists, %s will be passed" % plot_corr.__name__.upper())
        # return
    data, _ = load_data.data(feature_file)
    label = load_data.get_label(label_file)
    data_df = pd.DataFrame()
    data_df_same = pd.DataFrame()
    data_df_all = pd.DataFrame()

    for i in range(len(label)):
        current_corr = []
        current_corr_same = []
        tmp_file = '%s/cache/box/%s_%s_%s.txt' % (root_path, cell_line, label[i], bin_number)
        tmp_file_same = '%s/cache/box/same_%s_%s_%s.txt' % (root_path, cell_line, label[i], bin_number)
        if not os.path.isfile(tmp_file) and not os.path.isfile(tmp_file_same):
            print("开始计算不同类数据")
            print("计算", label[i], "")
            data_y = data[0:tad_num, i * feature_dimensional:(i + 1) * feature_dimensional]
            data_n = data[tad_num:data.shape[0], i * feature_dimensional:(i + 1) * feature_dimensional]
            data_mean = np.mean(np.vstack((data_y, data_n)))
            data_y = data_y - data_mean
            data_n = data_n - data_mean
            for m in data_y:
                m = np.array(m).reshape(feature_dimensional, )
                for n in data_n:
                    n = np.array(n).reshape(feature_dimensional, )
                    if np.linalg.norm(m) * np.linalg.norm(n) != 0:
                        tmp_corr = np.dot(m, n) / (np.linalg.norm(m) * np.linalg.norm(n))
                        current_corr.append(tmp_corr)
            np.savetxt(tmp_file, np.array(current_corr), fmt='%0.3f')
            print("计算", label[i], "结束")
        else:
            with open(tmp_file) as f:
                for line in f.readlines():
                    current_corr.append(float(line))
        if not os.path.isfile(tmp_file_same):
            print("开始计算同类数据")
            print("计算", label[i], "")
            data_y = data[0:tad_num, i * feature_dimensional:(i + 1) * feature_dimensional]
            data_n = data[tad_num:data.shape[0], i * feature_dimensional:(i + 1) * feature_dimensional]
            # print(np.mean(data_y))
            # print(np.mean(data_n))
            # data_y_mean = np.mean(data_y)
            # data_n_mean = np.mean(data_n)
            # data_y = data_y - data_y_mean
            # data_n = data_n - data_n_mean
            tmp_row = 0
            for m in data_y:
                tmp_row += 1
                m = np.array(m).reshape(feature_dimensional, )
                if tmp_row < data_y.shape[0] - 1:
                    for n in data_y[tmp_row:]:
                        n = np.array(n).reshape(feature_dimensional, )
                        if np.linalg.norm(m) * np.linalg.norm(n) != 0:
                            tmp_corr = np.dot(m, n) / (np.linalg.norm(m) * np.linalg.norm(n))
                            current_corr_same.append(tmp_corr)
            tmp_row = 0
            for m in data_n:
                tmp_row += 1
                m = np.array(m).reshape(feature_dimensional, )
                if tmp_row < data_n.shape[0] - 1:
                    for n in data_n[tmp_row:]:
                        n = np.array(n).reshape(feature_dimensional, )
                        if np.linalg.norm(m) * np.linalg.norm(n) != 0:
                            tmp_corr = np.dot(m, n) / (np.linalg.norm(m) * np.linalg.norm(n))
                            current_corr_same.append(tmp_corr)
            np.savetxt(tmp_file_same, np.array(current_corr_same), fmt='%0.3f')
            print("计算", label[i], "结束")
        else:
            with open(tmp_file_same) as f:
                for line in f.readlines():
                    current_corr_same.append(float(line))
        tmp_des = pd.Series(current_corr).describe()
        one = tmp_des.iloc[4]
        two = tmp_des.iloc[5]
        three = tmp_des.iloc[6]
        in_count = 0
        for j in current_corr:
            if j < one or j > three:
                current_corr[in_count] = two
            in_count += 1
        data_df.insert(i, label[i], pd.Series(current_corr).drop_duplicates())
        data_df_all.insert(i, label[i] + '_intra', pd.Series(current_corr).drop_duplicates())
        tmp_des_same = pd.Series(current_corr_same).describe()
        one = tmp_des_same.iloc[4]
        two = tmp_des_same.iloc[5]
        three = tmp_des_same.iloc[6]
        in_count = 0
        for j in current_corr_same:
            if j < one or j > three:
                current_corr_same[in_count] = two
            in_count += 1
        data_df_same.insert(i, label[i], pd.Series(current_corr_same).drop_duplicates())
        data_df_all.insert(i, label[i] + '_inter', pd.Series(current_corr_same).drop_duplicates())
    # print(data_df.describe().T)
    # print(data_df_same.describe().T)
    data_df_all.to_csv('%s/box/all.csv' % pic_path)
    data_df.boxplot(fontsize=7, grid=False)
    plt.savefig('%s/box/%s_%s_coscorref.eps' % (pic_path, cell_line, bin_number), format='eps', dpi=dpi)
    plt.close()
    data_df_same.boxplot(fontsize=7, grid=False)
    plt.savefig('%s/box/%s_%s_same_coscorref.eps' % (pic_path, cell_line, bin_number), format='eps', dpi=dpi)
    plt.close()
    fig, ax = plt.subplots()
    data_df_all.boxplot(fontsize=4, grid=False)
    fig.savefig('%s/box/%s_%s_all_coscorref.eps' % (pic_path, cell_line, bin_number), format='eps', dpi=dpi)


dpi = 100
tad_num = 2208
tad_not_num = 2208
num_classes = 2
bin_size = 40000
bin_number = 10
histone_type = 9
cell_type = "E017"
feature_dimensional = bin_number * 2 + 1
root_path = "/home/loopgan/workspace/Python/TAD-Lactuca"
pic_path = "%s/pic/%s" % (root_path, cell_type)
interval_dir = "%s/cache/%s/down_up_bin_num_%s_%skb/" % (
    root_path, cell_type, bin_number, int(int(bin_number * bin_size) / 1000))
feature_file = "%s/cache/%s/feature/bin_%s_%skb.xlsx" % (
    root_path, cell_type, bin_number, int((bin_size * bin_number) / 1000))
label_file = "%s/data/feature_index.txt" % root_path
cell_type_dir = '/home/loopgan/data/TAD/%s/' % cell_type

if __name__ == '__main__':
    for i in ["E017_new"]:
        if not os.path.exists(interval_dir):
            os.mkdir(interval_dir)
            # get_interval_different_resolution(interval_dir=interval_dir, bin_size=bin_size, bin_number=bin_number)
        # get_interval_signal(interval_dir=interval_dir, cell_type_dir=cell_type_dir.replace("E017", i))
        # write_to_xlsx(signal_file_dir=interval_dir, feature_file=feature_file)
        plot_dis(feature_file=feature_file, label_file=label_file, cell_line=cell_type,
                 feature_dimensional=feature_dimensional)
        # plot_corr(feature_file=feature_file, label_file=label_file, cell_line=i)
